from odoo import models, fields, api
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
import base64
from openpyxl.styles import Alignment
from io import BytesIO
from openpyxl.styles import Font
from odoo.http import request
from odoo.exceptions import ValidationError, AccessError, UserError
from collections import defaultdict
from datetime import date, timedelta, timezone
from pytz import timezone, UTC
from datetime import datetime
from calendar import monthrange
from openpyxl.styles import NamedStyle
import pytz

class banreservas(models.Model):
    _name = 'banreservas.forms'
    _description = 'Formularios para banreservas'

    campaign_id = fields.Char(related="partner_id.campaign", string="Campaign", store=True)
    campaign_month = fields.Char(related="partner_id.campaing_month", string="Mes de la campaña", store=True)

    assigned_user_id = fields.Many2one(related="partner_id.user_id", string="Assigned User", store=True, readonly=False)
    client_phone = fields.Char(related="partner_id.phone", string="Telefono", store=True)
    client_phone_2 = fields.Char(related="partner_id.phone_2", string="Telefono 2", store=True)
    client_phone_3 = fields.Char(related="partner_id.phone_3", string="Telefono 3", store=True)
    client_phone_4 = fields.Char(related="partner_id.phone_4", string="Telefono 4", store=True)
    client_phone_5 = fields.Char(related="partner_id.phone_5", string="Telefono 5", store=True)
    client_phone_6 = fields.Char(related="partner_id.phone_6", string="Telefono 6", store=True)
    client_mobile = fields.Char(related="partner_id.mobile", string="Celular", store=True)
    client_mobile_2 = fields.Char(related="partner_id.mobile_2", string="Celular 2", store=True)
    client_mobile_3 = fields.Char(related="partner_id.mobile_3", string="Celular 3", store=True)
    client_mobile_4 = fields.Char(related="partner_id.mobile_4", string="Celular 4", store=True)
    client_mobile_5 = fields.Char(related="partner_id.mobile_5", string="Celular 5", store=True)
    client_mobile_6 = fields.Char(related="partner_id.mobile_6", string="Celular 6", store=True)
    
    form_state_section = fields.Selection([
        ('sale', 'Venta'),
        ('not_instreted', 'No Interesado'),
        ('incomplete', 'Incompleto'),
        ('pending', 'Pendiente')
    ], string="Estado de la Sección",required=True)
    

    form_state = fields.Selection([
        ('sale', 'Venta'),
        ('economic_issues', 'Problemas Económicos'),
        ('limit_increase', 'Ya Tiene El Producto, Desea Aumento Límite'),
        ('already_has_product', 'Ya Tiene El Producto'),
        ('too_many_credit_cards', 'Tiene Muchas Tarjetas De Crédito'),
        ('not_interested', 'No Interesado'),
        ('income_below_10000', 'No Aplica/Ingresos Menores De $10,000'),
        ('age_range_exceeded', 'No Aplica/Fuera De Rango De Edad'),
        ('wants_to_visit_office', 'No Interesado/Desea Pasar Por La Oficina'),
        ('dissatisfied', 'No Interesado/Disgustado Con La Institución'),
        ('not_applicable_unemployed', 'No Aplica/No Está Laborando'),
        ('less_than_6_months', 'No Aplica/Menos De 6 Meses Laborando'),
        ('foreigner', 'No Aplica/Extranjero'),
        ('cash_only', 'No Interesado/Se Maneja Con Efectivo'),
        ('not_interested_loan', 'No Interesado/Desea Préstamo'),
        ('wants_amount_info', 'No Interesado/Desea Saber Monto Para Que Aplica'),
        ('banreservas_employee', 'No Aplica/Empleado De Banreservas'),
        ('credit_recovery', 'Recuperación De Crédito'),
        ('disconnected', 'Colgó'),
        ('very_low_limits', 'Límites Muy Bajos'),
        ('wrong_phone', 'Teléfono Equivocado'),
        ('dont_call_the_client', 'No Llamar Al Cliente'),
        ('outside_country', 'Fuera Del País'),
        ('died', 'Fallecido'),
        ('answering_machine', 'Máquina Contestadora'),
        ('busy', 'Ocupado'),
        ('no_answer', 'No Contestan'),
        ('audio_problems', 'Problemas De Audio'),
        ('call_later', 'Llamar Luego'),
        ('not_service','Teléfono Fuera De Servicio'),
        ('not_residential_contact','No Aplica/No Posee Contacto Residencial')
    ],string="Disposiciones",required=True)

    close_date = fields.Date(string="Fecha de cierre", required=True)

    #PRODUCTO
    product_type = fields.Many2one(
        'res.banreservas.cards',
        string="Tipo de producto",
        ondelete="set null")
        
    master_card = fields.Boolean(string="MasterCard", default=True)
    requested_amount = fields.Monetary(string="Monto Solicitado RD")
    requested_amount_US = fields.Monetary(string="Monto Solicitado US")
    names = fields.Char(string='Nombres')
    lastnames = fields.Char(string='Apellidos')

    document = fields.Char(string='Cédula de identidad',required=True)
    marital_status = fields.Selection([
        ('single', 'Soltero/a'),
        ('married', 'Casado/a')
    ], string="Estado Civil")
    gender = fields.Selection([
        ('male', 'Masculino'),
        ('female', 'Femenino')
    ], string="Sexo")
    nationality = fields.Char(string='Nacionalidad')
    birth_date = fields.Date(string='Fecha de nacimiento')
    birth_place = fields.Char(string='Lugar de nacimiento')

    #address information
    address_information = fields.Selection([
        ('home', 'Casa'),
        ('apartment', 'Apartamento')
    ], string='Datos de domicilio')
    type_of_housing = fields.Selection([
        ('rent', 'Alquilada'),
        ('own', 'Propia')
    ], string='Tipo de vivienda')
    street = fields.Char(string='Calle')

    street_number = fields.Integer(string='Numero')
    reference = fields.Char(string='Referencia')
    building = fields.Char(string='Edificio')
    building_number = fields.Char(string='Número de apartamento')
    sector = fields.Char(string='Sector')
    state = fields.Many2one('res.country.state', string='Provincia')
    municipality = fields.Char(string='Municipio')
    borough = fields.Char(string='Distrito Municipal')
    residential_phone = fields.Char(string='Teléfono residencial')
    cellphone = fields.Char(string="Celular")
    telephone_company = fields.Char(string='Compañia de telefono')
    email = fields.Char(string= "Correo")

    guarantee = fields.Selection([
        ('solidary', 'Solidaria'),
    ], string='Garantia')
    guarantee_value = fields.Monetary(string="Valor")
    description = fields.Text(string='Description de garantia')
    economic_activity = fields.Char(string='Actividad economica')

    # datos del conyugue
    spouse_names = fields.Char(string='Nombre(s)')
    spouse_lastnames = fields.Char(string='Apellido(s)')
    spouse_document = fields.Char(string='Número de cédula o pasaporte')

    # work data
    work_type = fields.Selection([
        ('worker', 'Empleado'),
        ('independent', 'Independiente'),
        ('local', 'Con local'),
        ('ambulant', 'Ambulante')
    ], string=' Tipo de trabajo')

    company_name = fields.Char(string='Nombre de la empresa')
    entry_date = fields.Date(string='Fecha de ingreso')
    occupation = fields.Char(string='Ocupacion')
    company_economic_activity = fields.Char(string='Actividad economica')
    employees_number = fields.Integer(string='Cantidad de empleados')
    monthly_income = fields.Monetary(string='Ingreso mensual RD$', currency_field='currency_id')

    currency_id = fields.Many2one('res.currency', string="Moneda", required=True, default=lambda self: self.env.company.currency_id)

    job_position = fields.Char(string='Posición que ocupa')
    supervisor_name = fields.Char(string='Nombre del supervisor inmediato')
    job_phone = fields.Char(string='Telefono')
    others_incomes = fields.Monetary(string='Otros Ingresos', currency_field='currency_id')
    others_incomes_info = fields.Char(string='Especifique')

    # company address

    company_street = fields.Char(string='Calle')
    company_street_number = fields.Integer(string='Numero')
    company_building = fields.Char(string='Edificio')
    company_apartment_number = fields.Integer(string='Numero de apartamento')
    company_sector = fields.Char(string='Sector')
    company_state = fields.Many2one('res.country.state', string='Provincia')
    company_municipality = fields.Char(string='Municipio')
    company_municipality_district = fields.Char(string='Distrito municipal')
    company_phone = fields.Char(string='Telefono')
    company_mail = fields.Char(string='Correo')

    # pariente cercano que no viva con usted

    relative_name = fields.Char(string='Nombre(s) y apellido(s)')
    kinship = fields.Char(string='Parentesco')
    relative_phones = fields.Char(string='Teléfono(s)')

    # referencias comerciales y personales

    name_1 = fields.Char(string='Nombre completo')
    adress_1 = fields.Char(string='Dirección')
    phone_1 = fields.Char(string='Telefono')

    # datos de la tarjeta

    payment_method = fields.Selection([
        ('total', 'Total'),
        ('minimun', 'Mínimo',)
    ])
    state_address = fields.Selection([
        ('residential', 'Residencia'),
        ('laboral', 'Laboral'),
        ('mail', 'Correo electrónico')
    ])
    card_address = fields.Many2one(
        'res.banreservas.offices',
        string="Oficina Comercial",
        ondelete="set null"
    )
    card_name = fields.Char(string='Como desea el nombre de la tarjeta')

    # tarjeta adicional
    additional_card = fields.Boolean(string="Tarjeta adicional")
    additional_card_names = fields.Char(string='Nombres')
    additional_card_lastnames = fields.Char(string='Apellidos')
    additional_card_document = fields.Char(string='Cédula de identidad / pasaporte')
    additional_card_kinship = fields.Char(string='Parentesco')
    additional_card_name = fields.Char(string='Como desea el nombre en la targeta')

    currency_id = fields.Many2one('res.currency', string="Moneda", default=lambda self: self.env.company.currency_id)

    Reason_not_applicable = fields.Char(string="Razon no aplicable")
    propose_limit_rd = fields.Char(string="Limite propuesto en RD", store=True)
    propose_limit_us = fields.Char(string="Limite propuesto en US")
    tc_propouse = fields.Char(string="TC propuesto")
    rd_amount_income_less_than_60m = fields.Monetary(string="Ingresos menores a 60m RD", currency_field='currency_id', related='partner_id.rd_amount_income_less_than_60m')
    tc_amount_income_less_than_60m = fields.Char(string="TC para ingresos menores de 60m", related='partner_id.tc_amount_income_less_than_60m')
    rd_amount_income_less_than_20m = fields.Monetary(string="Ingresos menores a 20m RD", currency_field='currency_id', related='partner_id.rd_amount_income_less_than_20m')
    tc_amount_income_less_than_20m = fields.Char(string="TC para ingresos menores de 20m", related='partner_id.tc_amount_income_less_than_20m')
    income_amount = fields.Monetary(string="Ingresos", currency_field='currency_id')
    edit = fields.Boolean(string='Desea editar los limites?')

    can_edit_limits = fields.Boolean(compute='_compute_can_edit_limits')

    comment = fields.Text(string="Comentario")
    partner_id = fields.Many2one(
        'res.partner',
        string="Cliente",
        domain="[('campaign', '=', 'banreservas')]", 
        store=True
    )
    _show_icons = {}

    def toggle_phone_icon(self):
        # Obtener el campo del número al que se aplica el cambio
        phone_field = self.env.context.get('phone_field')
        if phone_field:
            # Cambiar el estado del ícono
            if phone_field not in self._show_icons:
                self._show_icons[phone_field] = False  # Inicialmente no mostrar el ícono
            self._show_icons[phone_field] = not self._show_icons[phone_field]
            
    @api.constrains('close_date')
    def _check_close_date(self):
        for record in self:
            if record.close_date:
                # Verificar si la fecha es futura
                if record.close_date > fields.Date.today():
                    raise UserError("No puedes seleccionar una fecha de cierre futura.")
                
    @api.onchange('close_date')
    def _check_close_date_before(self):
        for record in self:
            if record.close_date and record.close_date < fields.Date.today():
                return {
                    'warning': {
                        'title': "Atención",
                        'message': f"Has ingresado una fecha de cierre pasada: {record.close_date}",
                    }
                }
                

    @api.onchange('edit')
    def _compute_can_edit_limits(self):
        group = self.env.ref('banreservas.group_edit_propose_limit')
        if group not in self.env.user.groups_id:
            self.can_edit_limits = False
        else:
            self.can_edit_limits = True
        
    @api.model
    def create(self, vals):
        record = super(banreservas, self).create(vals)
        record.partner_id._update_form_status()
        return record

    def write(self, vals):
        res = super(banreservas, self).write(vals)
        for record in self:
            record.partner_id._update_form_status()
        return res

    def unlink(self):
        partners = self.mapped('partner_id')
        res = super(banreservas, self).unlink()
        for partner in partners:
            partner._update_form_status()
        return res
    
    def change_phone(self):
        for record in self:
            record.client_phone = record.partner_id.phone
            record.client_phone_2 = record.partner_id.phone_2
            record.client_phone_3 = record.partner_id.phone_3
            record.client_phone_4 = record.partner_id.phone_4
            record.client_phone_5 = record.partner_id.phone_5
            record.client_phone_6 = record.partner_id.phone_6
            record.client_mobile = record.partner_id.mobile
            record.client_mobile_2 = record.partner_id.mobile_2
            record.client_mobile_3 = record.partner_id.mobile_3
            record.client_mobile_4 = record.partner_id.mobile_4
            record.client_mobile_5 = record.partner_id.mobile_5
            record.client_mobile_6 = record.partner_id.mobile_6
    

    def change_phone_2(self):
        for record in self:
            record.client_phone_2 = record.partner_id.phone_5 or ''

    def change_phone_3(self):
        for record in self:
            record.client_phone_3 = record.partner_id.phone_6 or ''

    def change_mobile(self):
        for record in self:
            record.client_mobile = record.partner_id.mobile_4 or ''

    def change_mobile_2(self):
        for record in self:
            record.client_mobile_2 = record.partner_id.mobile_5 or ''

    def change_mobile_3(self):
        for record in self:
            record.client_mobile_3 = record.partner_id.mobile_6 or ''

    
    def assign_existing_forms_to_clients(self):
        """
        Asocia los formularios existentes a los clientes correspondientes
        y actualiza los campos `form_id` y `has_form` en `res.partner`.
        """
        partners = self.env['res.partner'].search([])
        for partner in partners:
            form = self.search([('partner_id', '=', partner.id)], limit=1)
            partner.form_id = form.id if form else False
            partner.is_gestioned = bool(form)

    def assign_existing_forms_to_clients(self):
        # Buscar todos los registros de formularios donde no se ha asignado un partner_id
        forms = self.env['banreservas.forms'].search([('partner_id', '=', False)])
        
        for form in forms:
            # Buscar el partner basado en el documento u otros criterios
            contact_id = self.env['res.partner'].search([
                ('document', '=', form.document),
                ('campaign', '=', 'banreservas')], limit=1)
            
            # Si se encuentra el partner, asignarlo al formulario
            if contact_id:
                form.partner_id = contact_id
 

    @api.constrains('form_state', 'form_state_section')
    def _check_form_state_valid(self):
        """Valida que el valor de 'form_state' sea válido según la sección seleccionada."""
        # Diccionario con las opciones válidas según la sección
        valid_states = {
            'sale': ['sale'],
            'not_instreted': ['economic_issues', 'limit_increase', 'already_has_product',
                'has_many_credit_cards', 'not_interested', 'income_below_10000',
                'age_range_exceeded', 
                'wants_to_visit_office', 'dissatisfied',
                'not_applicable_unemployed', 'less_than_6_months',
                'foreigner', 'cash_only', 'not_interested_loan',
                'wants_amount_info', 'banreservas_employee',
                'credit_recovery', 'disconnected', 'very_low_limits', 'too_many_credit_cards'],
            'incomplete': ['wrong_phone', 'dont_call_the_client', 'outside_country', 'not_service', 'died'],
            'pending': ['answering_machine', 'busy', 'no_answer', 'audio_problems', 'call_later']
        }

        section_names = dict(self._fields['form_state_section'].selection)
        state_names = dict(self._fields['form_state'].selection)

        if self.form_state and self.form_state not in valid_states.get(self.form_state_section, []):
            # Obtener los nombres legibles para la sección y el estado
            section_name = section_names.get(self.form_state_section, self.form_state_section)
            state_name = state_names.get(self.form_state, self.form_state)

            raise UserError(
                f"El valor '{state_name}' no es válido para la sección '{section_name}'."
            )

    @api.model
    def calculate_daily_forms_for_banreservas(self):
        # Buscar formularios con estado 'sale' y agrupados por fecha
        forms = self.search([('form_state_section', '=', 'sale')])

        # Inicializar un diccionario para almacenar la cantidad de formularios por día
        daily_totals = defaultdict(int)

        # Recorrer los formularios y contarlos por fecha
        for form in forms:
            # Obtener la fecha (solo la parte de la fecha, sin hora)
            form_date = form.close_date.date()

            # Contar el formulario por la fecha
            daily_totals[form_date] += 1

        return daily_totals 
            

    def get_form_state_section_counts(self):
        """
        Genera un diccionario con el conteo de registros agrupados por form_state_section y form_state.
        """
        # Diccionario para almacenar los conteos
        result_dict = defaultdict(lambda: defaultdict(int))

        # Obtener todos los registros de 'banreservas.forms'
        forms = self.env['banreservas.forms'].search([])

        # Recorrer los formularios y agrupar por form_state_section y form_state
        for form in forms:
            section = form.form_state_section
            state = form.form_state
            result_dict[section][state] += 1

        # Traducir las claves (opcional: si las selecciones tienen etiquetas legibles)
        final_dict = {}
        section_labels = dict(self._fields['form_state_section'].selection)
        state_labels = dict(self._fields['form_state'].selection)

        for section, states in result_dict.items():
            section_label = section_labels.get(section, section)  # Usar etiqueta o valor crudo
            final_dict[section_label] = {}
            for state, count in states.items():
                state_label = state_labels.get(state, state)
                final_dict[section_label][state_label] = count

        return final_dict

    def get_form_state_count_grouped_by_day(self):
        # Inicializar un diccionario para almacenar los conteos por fecha
        count_dict = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

        # Filtrar todos los formularios
        forms = self.env['banreservas.forms'].search([])

        # Contar los registros por fecha, 'form_state_section' y 'form_state'
        for form in forms:
            form_date = form.close_date.date()  # Tomar la fecha de creación del formulario
            section = form.form_state_section or 'Sin Sección'  # Manejar valores vacíos
            state = form.form_state or 'Sin Estado'  # Manejar valores vacíos
            count_dict[form_date][section][state] += 1

        # Generar resultados organizados por fecha
        result_dict = {}
        for date, sections in count_dict.items():
            result_dict[date] = {}
            for section, states in sections.items():
                # Obtener etiquetas amigables para las selecciones
                section_label = dict(self._fields['form_state_section'].selection).get(section, section)
                result_dict[date][section_label] = {}
                for state, count in states.items():
                    state_label = dict(self._fields['form_state'].selection).get(state, state)
                    result_dict[date][section_label][state_label] = count

        return result_dict


    @api.model
    def default_get(self, fields_list):
        res = super(banreservas, self).default_get(fields_list)
        
        contact_id = request.session.get('default_contact_id')
        # _logger.info('ID de contacto desde sesión: %s', contact_id)
        
        # Verificar y asignar el contact_id
        if contact_id:
            res.update({'partner_id': contact_id})
            # _logger.info('Valores predeterminados actualizados: %s', res)
        
        return res

    @api.onchange('income_amount')
    def _compute_income_amount(self):
        for record in self:
            contact = self.env['res.partner'].search([('document', '=', record.document)], limit=1)
            if contact:
                if 0 < record.income_amount <= 20000:
                    # Actualizar los campos límite propuesto y TC propuesto
                    record.propose_limit_rd = contact.rd_amount_income_less_than_20m if contact.rd_amount_income_less_than_20m else contact.propose_limit_rd
                    record.tc_propouse = contact.tc_amount_income_less_than_20m if contact.tc_amount_income_less_than_20m else contact.tc_propouse
                    record.propose_limit_us = ''

                elif 20000 < record.income_amount <= 60000:
                    # Actualizar los campos límite propuesto y TC propuesto
                    record.propose_limit_rd = contact.rd_amount_income_less_than_60m if contact.rd_amount_income_less_than_60m else contact.propose_limit_rd
                    record.tc_propouse = contact.tc_amount_income_less_than_60m if contact.tc_amount_income_less_than_60m else contact.tc_propouse
                    record.propose_limit_us = ''

    @api.onchange('document', 'partner_id')
    def _compute_data (self):
        for record in self:
                if not record.partner_id and record.document:
            # Buscar el contacto
                    contact_id = self.env['res.partner'].search([
                        ('document', '=', record.document), 
                        ('campaign', '=', 'banreservas')], limit=1)
                    if contact_id:
                        record.partner_id = contact_id
                if record.partner_id:
                    # record.partner_id = contact_id
                    record.names = record.partner_id.first_name.strip()
                    record.document = record.partner_id.document
                    record.lastnames = record.partner_id.last_name
                    record.propose_limit_rd = record.partner_id.propose_limit_rd
                    record.propose_limit_us = record.partner_id.propose_limit_us
                    record.tc_propouse = record.partner_id.tc_propouse
                    record.requested_amount = record.partner_id.propose_limit_rd
                    record.requested_amount_US = record.partner_id.propose_limit_us
                    #record.additional_card_names = record.partner_id.first_name.strip().split()[0]
                    #record.additional_card_lastnames = record.partner_id.last_name.split()[0]

    @api.onchange('document')
    def _check_document_unique(self):
        for record in self:
            if self.search_count([('document', '=', record.document)]) >= 1:
                raise ValidationError("Existe un formulario creado para este cliente")
            

    @api.model
    def get_forms_grouped_by_month_and_day(self,date):
        """
        Agrupa los formularios por mes y, dentro de cada mes, cuenta los formularios por día,
        devolviendo el nombre del día correspondiente.

        :return: Diccionario donde las claves son meses ('YYYY-MM') y los valores
                son listas de diccionarios con las fechas, el conteo y el nombre del día.
        """
        query = """
            SELECT TO_CHAR(close_date, 'YYYY-MM') AS month,
                TO_CHAR(close_date, 'YYYY-MM-DD') AS full_date,
                TO_CHAR(close_date, 'Day') AS day_name,
                COUNT(*) AS count
            FROM banreservas_forms
            WHERE close_date IS NOT NULL
        """

        if date:
            query += " AND close_date >= %s"
            params = date
        else:
            params = []

        query += """
                GROUP BY month, full_date, day_name
                ORDER BY month, full_date
    
            """

        self.env.cr.execute(query,(params,))
        results = self.env.cr.fetchall()

        # Crear un diccionario para agrupar los formularios
        grouped_forms = defaultdict(list)
        for month, full_date, day_name, count in results:
            grouped_forms[month].append({
                "date": full_date,
                "count": count,
                "day_name": day_name.strip()  # Eliminamos espacios adicionales
            })

        # Convertir el defaultdict a un diccionario estándar
        return dict(grouped_forms)

    def export_forms_excel(self,forms_date):
            plantilla_path = '/mnt/extra-addons/banreservas/views/Formulario TCR-084.xlsx'
            wb = load_workbook(plantilla_path,keep_vba=True)
            ws_template = wb.active
            fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            forms_date_dt = datetime.combine(forms_date, datetime.min.time())
            number_style = NamedStyle(name="number_style", number_format="00000000000")

            # Buscar todos los registros
            total_forms = self.search([
                ('form_state_section', '=', 'sale'),
                ('close_date', '>=', forms_date_dt),  # Asegúrate de que sea mayor o igual a forms_date con hora 00:00:00
                ('close_date', '<', forms_date_dt + timedelta(days=1))
                ])

            for form in total_forms:            
                # Crear una nueva hoja para cada registro
                ws = wb.copy_worksheet(ws_template)
                for row in ws.iter_rows(min_row=1, max_row=8, min_col=1, max_col=33):
                            for cell in row:
                                cell.fill = fill_white
                for row in ws.iter_rows(min_col=1, max_col=1): 
                    for cell in row:
                        cell.fill = fill_white

                ws.title = f"{form.document}"

                for image in ws_template._images:
                    img_copy = Image('/mnt/extra-addons/banreservas/views/Banreservas_logo.png')  # image.path debe contener la ruta de la imagen
                    img_copy.anchor = image.anchor 
                    ws.add_image(img_copy) 

                data = {
                    "E9": "MasterCard " + str(form.tc_propouse),
                    "AA9":int(float((form.propose_limit_rd))) if form.propose_limit_rd else 0,
                    "AA10":int(float((form.propose_limit_us))) if form.propose_limit_us else 0,
                    "E14": form.names.strip(),
                    "U14": form.lastnames,
                    "AA16": int(form.document),
                    "D24": form.street,
                    "D26": form.reference,
                    "D28": form.building,
                    "D30": form.sector,
                    "D32": form.municipality,
                    "F34": form.residential_phone,
                    "F36": form.email,
                    "N34": form.cellphone,
                    "AD24": form.street_number,
                    "X28": form.building_number,
                    "X30": form.state.name,
                    "X32": form.borough,
                    "X34": form.telephone_company,
                    "P38": form.guarantee_value,
                    "X38": form.description,
                    "F41": form.economic_activity,
                    "E46": form.spouse_names,
                    "G48": form.spouse_document,
                    "X46": form.spouse_lastnames,
                    "F52": form.company_name,
                    "F54": form.company_economic_activity,
                    "F56": form.employees_number,
                    "F58": form.job_position,
                    "F60": form.job_phone,
                    "F62": form.others_incomes_info,
                    "X52": form.entry_date,
                    "X54": form.occupation,
                    "X56": form.monthly_income,
                    "Z58": form.supervisor_name,
                    "X60": form.others_incomes,
                    "E65": form.company_street,
                    "E67": form.company_building,
                    "E69": form.company_sector,
                    "E71": form.company_municipality,
                    "E73": form.company_phone,
                    "AD65": form.company_street_number,
                    "X67": form.company_apartment_number,
                    "X69": form.company_state.name,
                    "X71": form.company_municipality_district,
                    "X73": form.company_mail,
                    "F77": form.relative_name,
                    "E79": form.kinship,
                    "X79": form.relative_phones,
                    "B84": form.name_1,
                    "G84": form.adress_1,
                    "Z84": form.phone_1,
                    "E100": form.additional_card_names,
                    "G102": int(form.additional_card_document),
                    "H104": form.additional_card_name,
                    "U100": form.additional_card_lastnames,
                    "U102": form.additional_card_kinship,
                    "E121": "4",
                    "R10": "Visa",
                    "S10": "☐ Ser 1",
                    "U10": "☐ Ser 2",
                    "W10": "☐ Ser 3",
                    "Y22": "Propia",
                    "AA50": "Con local",
                    "V111": form.create_uid.br_code.replace("-", ""),
                    "Z93": form.card_address.code,
                    "P18": form.nationality,
                    "AA18": form.birth_date,
                    "AA20": form.birth_place,
                    "AA7": form.close_date,
                    "F117": "MasterCard " + str(form.tc_propouse),
                    "B121": f"US$ {int(float((form.propose_limit_us))) if form.propose_limit_us else 0}",
                    "H96": f"{(form.names.split()[0] if form.names else '')} {(form.lastnames.split()[0] if form.lastnames else '')}"
                }

                cells_to_format = [
                    'G38', 'E38', 'Z50', 'AC50', 'D18', 'D20', 'H18', 'H20', 
                    'H22', 'L22', 'X22', 'AB22', 'E38', 'G38', 'G50', 'R50', 
                    'Z50', 'AC50', 'H89', 'O89', 'H91', 'O91', 'T91', 'H93', 
                    'O93', 'T93', 'M98', 'I98','Q9','Q10','H16','Q16', 'AA16'
                ]

                font = Font(size=13)

                # Aplicación de estilos a cada celda en la lista
                for cell in cells_to_format:
                    if cell == 'AA16':
                        ws[cell].style = number_style
                        ws[cell].alignment = Alignment(horizontal='left')

                    if cell != 'AA16':
                        ws[cell].font = font
                        ws[cell].alignment = Alignment(horizontal='right',vertical='top')
                      
                

                ws['Q9'] = "🗹" if form.master_card == True else "☐"
                ws['Q10'] = "☐"
                ws['B119'] = f"RD$ {int(float((form.propose_limit_rd))) if form.propose_limit_rd else 0}"

                ws['H16'] = "🗹"
                ws['Q16'] = "☐"

                if form.marital_status == "married":
                    ws['D18'] = "🗹"
                    ws['D20'] = "☐"

                elif form.marital_status == "single":
                    ws['D20'] = "🗹"
                    ws['D18'] = "☐"
                else:
                    ws['D20'] = "☐"
                    ws['D18'] = "☐"

                if form.gender == "male":
                    ws['H18'] = "🗹"
                    ws['H20'] = "☐"
                elif form.gender == "female":
                    ws['H18'] = "☐"
                    ws['H20'] = "🗹"
                else:
                    ws['H18'] = "☐"
                    ws['H20'] = "☐"

                if form.address_information == "home":
                    ws['H22'] = "🗹"
                    ws['L22'] = '☐'
                elif form.address_information == "apartment": 
                    ws['L22'] = "🗹"
                    ws['H22'] = '☐'
                else:
                    ws['L22'] = "☐"
                    ws['H22'] = '☐'

                if form.type_of_housing == 'own':
                    ws['X22'] = "🗹"
                    ws['AB22'] = '☐'
                elif form.type_of_housing == "rent":
                    ws['AB22'] = "🗹"
                    ws['X22'] = '☐'
                else:
                    ws['AB22'] = "☐"
                    ws['X22'] = '☐'
                
                if form.guarantee == 'solidary':
                    ws['E38'] = "🗹"
                    ws['G38'] = '☐'
                else:
                    ws['G38'] = "☐"
                    ws['E38'] = '☐'

                if form.work_type == "worker":
                    ws['G50'] = "🗹"
                    ws['R50'] = '☐'
                    ws['Z50'] = "☐"
                    ws['AC50'] = '☐'
                elif form.work_type == "independent":
                    ws['G50'] = "☐"
                    ws['R50'] = '🗹'
                    ws['Z50'] = "☐"
                    ws['AC50'] = '☐'
                elif form.work_type == "local":
                    ws['G50'] = "☐"
                    ws['R50'] = '☐'
                    ws['Z50'] = "🗹"
                    ws['AC50'] = '☐'
                elif form.work_type == "ambulant":
                    ws['G50'] = "☐"
                    ws['R50'] = '☐'
                    ws['Z50'] = "☐"
                    ws['AC50'] = '🗹'
                else:
                    ws['G50'] = "☐"
                    ws['R50'] = '☐'
                    ws['Z50'] = "☐"
                    ws['AC50'] = '☐'

                if form.additional_card == True:
                    ws['I98'] = '🗹'
                    ws['M98'] = "☐"
                else:
                    ws['I98'] = "☐"
                    ws['M98'] = '🗹'


                ws['O89'] = "☐"
                ws['H89'] = '☐'
                ws['H91'] = "☐"
                ws['O91'] = '☐'
                ws['T91'] = '☐'
                ws['H93'] = "☐"
                ws['O93'] = '☐'
                ws['T93'] = '🗹'

                for cell, value in data.items():
                    ws[cell] = value if value else ""
                    ws[cell].font = Font(color="000000")
                    
                        

            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)

            # Codificar el archivo en base64 para que Odoo pueda manejarlo como binario
            archivo_excel = base64.b64encode(fp.read())
            
            fp.close()

            nombre_archivo = 'formulario_clientes.csv'
            attachment = self.env['ir.attachment'].create({
                'name': nombre_archivo,
                'type': 'binary',
                'datas': archivo_excel,
                'store_fname': nombre_archivo,
                'res_model': 'formulario.cliente',
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })

            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'self',
            }

    def count_forms_by_day_and_state_current_month(self, state, date):
        query = """
            SELECT 
                DATE_TRUNC('day', close_date) AS day,
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                DATE_TRUNC('day', close_date) = %s
                AND form_state_section = %s
            GROUP BY 
                day
            ORDER BY 
                day;
        """
        self.env.cr.execute(query, (date, state,))
        results = self.env.cr.fetchall()

        if results:
            return results[0][1]  
        return 0 
   
    def total_forms_by_month_and_year(self, disposition, date):
        """
        Obtiene el conteo de formularios cerrados hasta el último día de un mes específico, incluyendo meses anteriores.

        :param disposition: Estado de la sección del formulario (form_state_section).
        :param date: Mes en formato 'YYYY-MM' para filtrar los formularios.
        :return: Conteo de formularios.
        """
        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                close_date <= (
                    DATE_TRUNC('month', TO_DATE(%s, 'YYYY-MM')) + INTERVAL '1 month' - INTERVAL '1 day'
                )
                AND form_state_section = %s
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (date, disposition))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0

    def total_pending_forms_by_month_and_year(self, date):
        """
        Obtiene la suma de los campos: answering_machine, busy, no_answer, audio_problems, call_later
        para una fecha específica.
        """
        query = """
            SELECT 
                SUM(answering_machine) AS total_answering_machine,
                SUM(busy) AS total_busy,
                SUM(no_answer) AS total_no_answer,
                SUM(audio_problems) AS total_audio_problems,
                SUM(call_later) AS total_call_later
            FROM 
                pending_forms
            WHERE 
                date <=  (
                    DATE_TRUNC('month', TO_DATE(%s, 'YYYY-MM')) + INTERVAL '1 month' - INTERVAL '1 day'
                )
        """
        
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (date,))
        result = self.env.cr.fetchone()

        # Manejar valores None y sumar los resultados
        total_sum = sum(value if value is not None else 0 for value in result) if result else 0
        return total_sum
            
    #ESTE NO VA
    def total_worked_leads_by_month_and_year(self, date):

        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                close_date <= (
                    DATE_TRUNC('month', TO_DATE(%s, 'YYYY-MM')) + INTERVAL '1 month' - INTERVAL '1 day'
                )
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (date,))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0

    def total_efective_contacts_by_year(self, date):

        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                close_date <= (
                    DATE_TRUNC('month', TO_DATE(%s, 'YYYY-MM')) + INTERVAL '1 month' - INTERVAL '1 day'
                )
                AND form_state_section = 'sale'
                OR form_state_section = 'not_instreted'
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (date,))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0
    
    def total_efective_contacts_by_month(self, date):
        year,month = date.split('-')

        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                EXTRACT(YEAR FROM close_date) = %s
                AND EXTRACT(MONTH FROM close_date) = %s
                AND form_state_section = 'sale'
                OR form_state_section = 'not_instreted'
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (year,month,))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0
    
    def total_worked_leads_by_month(self, date):
        year,month = date.split('-')

        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                EXTRACT(YEAR FROM close_date) = %s
                AND EXTRACT(MONTH FROM close_date) = %s
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (year,month,))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0

    #ESTE NO VA
    def total_forms_by_year_and_disposition(self,date,state,tc_propouse):
        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                close_date <= (
                    DATE_TRUNC('month', TO_DATE(%s, 'YYYY-MM')) + INTERVAL '1 month' - INTERVAL '1 day'
                )
                AND form_state = %s
                AND tc_propouse = %s
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (date, state,tc_propouse))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0
    
    def total_forms_by_month(self, disposition, date):
        year, month = date.split('-')

        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                EXTRACT(YEAR FROM close_date) = %s
                AND EXTRACT(MONTH FROM close_date) = %s
                AND form_state_section = %s
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (year,month, disposition))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0
    
    #ESTE NO VA
    def count_total_forms_by_disposition(self, date, state, section):
        year, month = date.split('-')

        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                EXTRACT(YEAR FROM close_date) = %s
                AND EXTRACT(MONTH FROM close_date) = %s
                AND form_state = %s
                AND form_state_section = %s
        """
        self.env.cr.execute(query, (int(year), int(month), state, section))
        results = self.env.cr.fetchone()

        return results[0] if results else 0

    def total_forms_by_year_and_section(self,date,state,section):
        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                close_date <= (
                    DATE_TRUNC('month', TO_DATE(%s, 'YYYY-MM')) + INTERVAL '1 month' - INTERVAL '1 day'
                )
                AND form_state = %s
                AND form_state_section = %s
        """
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (date, state,section))
        result = self.env.cr.fetchone()

        # Retornar el conteo si hay resultados; de lo contrario, 0
        return result[0] if result else 0

    def total_pending_forms_by_year_and_section(self,date,disposition):
        query = """
            SELECT 
                SUM({0}) AS total_disposition
            FROM 
                pending_forms
            WHERE 
                date <= (
                    DATE_TRUNC('month', TO_DATE(%s, 'YYYY-MM')) + INTERVAL '1 month' - INTERVAL '1 day'
                )
        """.format(disposition)
        # Ejecutar la consulta SQL con parámetros dinámicos
        self.env.cr.execute(query, (date,))
        result = self.env.cr.fetchone()

        total_sum = sum(value if value is not None else 0 for value in result) if result else 0
        return total_sum
    
    def count_forms_by_disposition(self, state, date, section):
        query = """
            SELECT 
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                DATE_TRUNC('day', close_date) = %s
                AND form_state = %s
                AND form_state_section = %s;
        """
        self.env.cr.execute(query, (date, state, section))
        result = self.env.cr.fetchone()

        return result[0] if result else 0
    
    #ESTE NO VA
    def get_total_all_forms(self,date):
        query = """
            SELECT 
                DATE_TRUNC('day', close_date) AS day,
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                DATE_TRUNC('day', close_date) = %s
                AND tc_propouse IS NOT NULL
                AND form_state_section = %s

            GROUP BY 
                day
            ORDER BY 
                day;
        """
        self.env.cr.execute(query, (date,'not_instreted',))
        results = self.env.cr.fetchall()

        if results:
            return results[0][1]  
        return 0 

    def count_forms_by_state_and_cards(self, state, date,card):
        query = """
            SELECT 
                DATE_TRUNC('day', close_date) AS day,
                COUNT(*) AS form_count
            FROM 
                banreservas_forms
            WHERE 
                DATE_TRUNC('day', close_date) = %s
                AND form_state = %s
                AND tc_propouse = %s
            GROUP BY 
                day
            ORDER BY 
                day;
        """
        self.env.cr.execute(query, (date, state, card,))
        results = self.env.cr.fetchall()

        if results:
            return results[0][1]  
        return 0 
   
    def column_number_to_letter(sel,col_num):
        """Convierte un número de columna en una letra o combinación de letras de Excel."""
        col_letter = ''
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            col_letter = chr(remainder + 65) + col_letter
        return col_letter

    def total_pending_forms_for_contacts(self, date):
        """
        Retorna la suma de los valores de call_later, audio_problems y el conteo de formularios.
        """
        query = """
            SELECT
                SUM(COALESCE(call_later, 0)) AS total_call_later,
                SUM(COALESCE(audio_problems, 0)) AS total_audio_problems
            FROM pending_forms
            WHERE date = %s
        """

        self.env.cr.execute(query, (date,))
        result = self.env.cr.fetchone()  # fetchone porque esperamos una sola fila

        # Aseguramos que los valores no sean None y los sumamos
        total_call_later = result[0] if result and result[0] is not None else 0
        total_audio_problems = result[1] if result and result[1] is not None else 0

        return total_call_later + total_audio_problems

    
    
    def contact_total(self,date):
        """
        Retorna todos los contactos que iran el la casilla de contactados
        """
        query = """
            SELECT
                COUNT(*) AS form_count
            FROM banreservas_forms
            WHERE
                DATE_TRUNC('day', close_date) = %s
                AND form_state_section != 'pending'
                AND form_state NOT IN ('wrong_phone', 'outside_country', 'not_service', 'died');
        """

        self.env.cr.execute(query,(date,))
        result = self.env.cr.fetchone()
        return result[0] if result else 0


   
    def export_form_productivity(self,date):
            template_path = '/mnt/extra-addons/banreservas/views/Productividad.xlsx'
            wb = load_workbook(template_path,keep_vba=True)
            ws_template = wb.active
            today = datetime.today()

            year = today.year
            month = today.month
            end_day = monthrange(year, month)[1]  
            end_date = datetime(year, month, end_day).date()


            call_data = self.env['calls.and.contacts'].search([])

            start_column = 3
            start_column_cards = 3

            for month, days in self.get_forms_grouped_by_month_and_day(date).items():
                ws = wb.copy_worksheet(ws_template)
                ws.title = f"{month}"
                count = 0
                total_complete = 0

                for image in ws_template._images:
                    img_copy = Image('/mnt/extra-addons/banreservas/views/Banreservas_logo.png')
                    img_copy.anchor = image.anchor 
                    ws.add_image(img_copy) 

                year, month_number = map(int, month.split('-'))

                # Calcular el último día del mes
                last_day = monthrange(year, month_number)[1]  # Devuelve un tuple (primer_día, último_día)
                end_date = f"{month}-{last_day}" 

                for day_info in days:
                    ws['B3'] = f"Fecha De Inicio: {month}-1"
                    ws['B4'] = f"Fecha De Término: {end_date}"
                    ws['B1'] = "Proveedor: Precisa Group SRL"
                    ws['B5'] = f"Fecha De Actividad: {today.date()}"
                    ws['AJ2'] = f"{today.date()}"
                    
                    num = start_column + count
                    col_letter = self.column_number_to_letter(num) 
                    col_letter_cards = self.column_number_to_letter(start_column_cards + count)
                    cell = f"{col_letter}6" 
                    date = day_info["date"]
                    day_name = day_info["day_name"]
                    additional_inf = call_data.get_totals_by_date(date)
                    ws[f"{col_letter_cards}15"] = round(additional_inf[0])
                    #CONTACTADOS
                    total_contacts = self.contact_total(date) + self.total_pending_forms_for_contacts(date)
                    ws[f"{col_letter_cards}18"] = total_contacts  #AQUIIIIIIIIIIIII
                     
                    total_complete += self.count_forms_by_day_and_state_current_month('sale',date)
                    ws['AH16'] = "0"
                    ws['AI16'] = total_complete
                    ws['AJ16'] = self.total_forms_by_year_and_section(date[:7],'sale','sale')
                    ws['AJ27'] = self.total_forms_by_month_and_year('sale',date[:7])
                    ws['AJ29'] = self.total_forms_by_month_and_year('not_instreted',date[:7])
                    ws['AJ53'] = self.total_forms_by_month_and_year('incomplete',date[:7])
                    ws['AJ64'] = self.total_pending_forms_by_month_and_year(date[:7])

                    #LEADS TRABAJADOS
                    worked_leads = self.env['banreservas.forms'].search_count([
                        ('close_date', '=', date),
                        ('form_state_section', '!=', 'pending')
                    ])
                    pending_total_month = self.env['pending.forms'].search([
                        ('date', '=', date)
                    ])
                    total_sum = sum(record.answering_machine + record.busy + record.no_answer + record.audio_problems + record.call_later for record in pending_total_month)
                    ws[f"{col_letter_cards}17"] = worked_leads + total_sum
                    
                    #CONTACTOS EFECTIVOS
                    ws[f"{col_letter_cards}19"] = len(
                        self.env['banreservas.forms'].search([
                            ('close_date', '>=', datetime.strptime(date, '%Y-%m-%d')),
                            ('close_date', '<', datetime.strptime(date, '%Y-%m-%d') + timedelta(days=1)),
                            '|',  
                            ('form_state_section', '=', 'sale'),
                            ('form_state_section', '=', 'not_instreted')
                        ])
                    )
                    #CONTACTABILIDAD
                    if worked_leads:
                        percentage = (total_contacts / (worked_leads + total_sum)) * 100
                        ws[f"{col_letter_cards}20"] = "{:.2f}%".format(percentage)  # Formatear como cadena con 2 decimales
                    else:
                        ws[f"{col_letter_cards}20"] = "0.00%"

                    ws['AJ19'] = self.total_efective_contacts_by_year(date[:7])
                    ws['AI19'] = self.total_efective_contacts_by_month(date[:7])

                    # Filtra los usuarios por la fecha de su último login
                    date_value = datetime.strptime(date, '%Y-%m-%d').date()

                    # Filtrar registros de inicio de sesión
  
                    #CONVERSION DATE
                    if total_contacts > 0:
                        ws[f'{col_letter}21'] = self.count_forms_by_day_and_state_current_month('sale',date) / total_contacts 
                    else:
                        ws[f'{col_letter}21'] = 0
                    #---------------------------------

                    # HOURS PER DAY
                    agent_work_model = self.env['agent.work.hours']
                    agent_hours_banreservas = self.env['agent.hours.banreservas']

                    # Buscar primero en el modelo agent.hours.banreservas
                    existing_record = agent_hours_banreservas.search([('date', '=', date)], limit=1)

                    if existing_record:
                        # Usar las horas del modelo si existe un registro
                        hours = existing_record.hours
                    else:
                        # Buscar en los resultados de la función si no hay registros
                        result = next((record for record in agent_work_model.calculate_daily_hours_for_banreservas() if record['date'] == date), None)
                        hours = result['hours'] if result else 0

                    # Asignar las horas al Excel
                    ws[f'{col_letter}8'] = hours

                    if hours > 0:
                        ws[f"{col_letter_cards}12"] = additional_inf[0] / hours
                        ws[f"{col_letter_cards}14"] = total_contacts / hours
                        ws[f"{col_letter}13"] = self.count_forms_by_day_and_state_current_month('sale', date) / hours
                        # SALES PER HOUR
                        ws[f"{col_letter}22"] = self.count_forms_by_day_and_state_current_month('sale', date) / hours
                    else:
                        ws[f"{col_letter}13"] = 0
                        # SALES PER HOUR
                        ws[f"{col_letter}22"] = 0

                    ws[f"{col_letter}16"] = self.count_forms_by_day_and_state_current_month('sale',date)
                    ws[cell] = day_name
                    ws[f"{col_letter}7"] = date
                    ws[f"{col_letter}27"] = self.count_forms_by_day_and_state_current_month('sale',date)
                    ws[f"{col_letter}29"] = self.count_forms_by_day_and_state_current_month('not_instreted',date)
                    ws[f"{col_letter}53"] = self.count_forms_by_day_and_state_current_month('incomplete',date)
                    ws[f"{col_letter}64"] = total_sum

                    #AGENTES LOGUEADOS
                    ws[f"{col_letter}23"] = additional_inf[3]


                    disposition_not_instreted = {
                        f'{col_letter}31':'economic_issues',
                        f'{col_letter}32':'limit_increase', 
                        f'{col_letter}33':'already_has_product', 
                        f'{col_letter}34':'too_many_credit_cards', 
                        f'{col_letter}35':'not_interested', 
                        f'{col_letter}36':'income_below_10000', 
                        f'{col_letter}37':'age_range_exceeded',
                        f'{col_letter}39':'wants_to_visit_office', 
                        f'{col_letter}40':'dissatisfied', 
                        f'{col_letter}41':'not_applicable_unemployed', 
                        f'{col_letter}42':'less_than_6_months', 
                        f'{col_letter}43':'foreigner',
                        f'{col_letter}44':'cash_only',
                        f'{col_letter}45':'not_interested_loan', 
                        f'{col_letter}46':'wants_amount_info',
                        f'{col_letter}47':'banreservas_employee', 
                        f'{col_letter}48':'credit_recovery', 
                        f'{col_letter}49':'disconnected', 
                        f'{col_letter}50':'very_low_limits'
                    }

                    disposition_pending = {
                        f'{col_letter}66': 'answering_machine',
                        f'{col_letter}67': 'busy',
                        f'{col_letter}68': 'no_answer',
                        f'{col_letter}69': 'audio_problems',
                        f'{col_letter}70': 'call_later'
                    }


                    disposition_incomplete = {
                        f'{col_letter}55': 'wrong_phone',
                        f'{col_letter}56': 'dont_call_the_client',
                        f'{col_letter}57': 'outside_country',
                        f'{col_letter}58': 'not_service',
                        f'{col_letter}59': 'died'
                    }

                    tz = pytz.timezone('America/Santo_Domingo')  # Ajusta según tu zona horaria
                    today_date = datetime.now(tz).date()

                    for key,value in disposition_not_instreted.items():
                        ws[key] = self.count_forms_by_disposition(value,date,'not_instreted')
                        if count == 0:
                            row_number = key[-2] + key[-1]
                            ws[f'AJ{row_number}'] = self.total_forms_by_year_and_section(date[:7],value,'not_instreted')

                    for key,value in disposition_pending.items():
                        pending_model = self.env['pending.forms'].search([
                            ('date', '=', date)
                        ])
                        pending = self.count_forms_by_disposition(value,date,'pending')

                        if pending_model:
                            formated_date = datetime.strptime(date, '%Y-%m-%d').date()
                            
                            
                            if formated_date == today_date:
                                pending_model.write({
                                value : pending if pending else 0
                                })
                        else:
                            self.env['pending.forms'].create({
                            value : pending if pending else 0,
                            'date': date,
                            })

                        for record in pending_model:
                            ws[key] = record[value]

                        if count == 0:
                            row_number = key[-2] + key[-1]
                            ws[f'AJ{row_number}'] = self.total_pending_forms_by_year_and_section(date[:7],value)


                    for key,value in disposition_incomplete.items():
                        ws[key] = self.count_forms_by_disposition(value,date,'incomplete')
                        if count == 0:
                            row_number = key[-2] + key[-1]
                            ws[f'AJ{row_number}'] = self.total_forms_by_year_and_section(date[:7],value,'incomplete')

                    clasic_dispositions = {
                        f'{col_letter_cards}77':'economic_issues',
                        f'{col_letter_cards}78':'limit_increase', 
                        f'{col_letter_cards}79':'already_has_product', 
                        f'{col_letter_cards}80':'too_many_credit_cards', 
                        f'{col_letter_cards}81':'not_interested', 
                        f'{col_letter_cards}82':'income_below_10000', 
                        f'{col_letter_cards}83':'age_range_exceeded',
                        f'{col_letter_cards}85':'wants_to_visit_office', 
                        f'{col_letter_cards}86':'dissatisfied', 
                        f'{col_letter_cards}87':'not_applicable_unemployed', 
                        f'{col_letter_cards}88':'less_than_6_months', 
                        f'{col_letter_cards}89':'foreigner',
                        f'{col_letter_cards}90':'cash_only',
                        f'{col_letter_cards}91':'not_interested_loan', 
                        f'{col_letter_cards}92':'wants_amount_info',
                        f'{col_letter_cards}93':'banreservas_employee', 
                        f'{col_letter_cards}94':'credit_recovery', 
                        f'{col_letter_cards}95':'disconnected', 
                        f'{col_letter_cards}96':'very_low_limits'
                    }

                    gold_dispositions = {
                        f'{col_letter_cards}98':'economic_issues',
                        f'{col_letter_cards}99':'limit_increase',
                        f'{col_letter_cards}100':'already_has_product',
                        f'{col_letter_cards}101':'too_many_credit_cards',
                        f'{col_letter_cards}102':'not_interested',
                        f'{col_letter_cards}103':'income_below_10000',
                        f'{col_letter_cards}104':'age_range_exceeded',
                        f'{col_letter_cards}106':'wants_to_visit_office',
                        f'{col_letter_cards}107':'dissatisfied',
                        f'{col_letter_cards}108':'not_applicable_unemployed',
                        f'{col_letter_cards}109':'less_than_6_months',
                        f'{col_letter_cards}110':'foreigner',
                        f'{col_letter_cards}111':'cash_only',
                        f'{col_letter_cards}112':'not_interested_loan',
                        f'{col_letter_cards}113':'wants_amount_info',
                        f'{col_letter_cards}114':'banreservas_employee',
                        f'{col_letter_cards}115':'credit_recovery',
                        f'{col_letter_cards}116':'disconnected',
                        f'{col_letter_cards}117':'very_low_limits'
                    }


                    multimoneda_dispositions = {
                        f'{col_letter_cards}119':'economic_issues',
                        f'{col_letter_cards}120':'limit_increase', 
                        f'{col_letter_cards}121':'already_has_product', 
                        f'{col_letter_cards}122':'too_many_credit_cards', 
                        f'{col_letter_cards}123':'not_interested', 
                        f'{col_letter_cards}124':'income_below_10000', 
                        f'{col_letter_cards}125':'age_range_exceeded',
                        f'{col_letter_cards}127':'wants_to_visit_office', 
                        f'{col_letter_cards}128':'dissatisfied', 
                        f'{col_letter_cards}129':'not_applicable_unemployed', 
                        f'{col_letter_cards}130':'less_than_6_months', 
                        f'{col_letter_cards}131':'foreigner',
                        f'{col_letter_cards}132':'cash_only',
                        f'{col_letter_cards}133':'not_interested_loan', 
                        f'{col_letter_cards}134':'wants_amount_info',
                        f'{col_letter_cards}135':'banreservas_employee', 
                        f'{col_letter_cards}136':'credit_recovery', 
                        f'{col_letter_cards}137':'disconnected', 
                        f'{col_letter_cards}138':'very_low_limits'
                    }


                    multimoneda_gold_dispositions = {
                        f'{col_letter_cards}140':'economic_issues',
                        f'{col_letter_cards}141':'limit_increase', 
                        f'{col_letter_cards}142':'already_has_product', 
                        f'{col_letter_cards}143':'too_many_credit_cards', 
                        f'{col_letter_cards}144':'not_interested', 
                        f'{col_letter_cards}145':'income_below_10000', 
                        f'{col_letter_cards}146':'age_range_exceeded',
                        f'{col_letter_cards}148':'wants_to_visit_office', 
                        f'{col_letter_cards}149':'dissatisfied', 
                        f'{col_letter_cards}150':'not_applicable_unemployed', 
                        f'{col_letter_cards}151':'less_than_6_months', 
                        f'{col_letter_cards}152':'foreigner',
                        f'{col_letter_cards}153':'cash_only',
                        f'{col_letter_cards}154':'not_interested_loan', 
                        f'{col_letter_cards}155':'wants_amount_info',
                        f'{col_letter_cards}156':'banreservas_employee', 
                        f'{col_letter_cards}157':'credit_recovery', 
                        f'{col_letter_cards}158':'disconnected', 
                        f'{col_letter_cards}159':'very_low_limits'
                    }

                    platinum_dispositions = {
                        f'{col_letter_cards}161':'economic_issues',
                        f'{col_letter_cards}162':'limit_increase', 
                        f'{col_letter_cards}163':'already_has_product', 
                        f'{col_letter_cards}164':'too_many_credit_cards', 
                        f'{col_letter_cards}165':'not_interested', 
                        f'{col_letter_cards}166':'income_below_10000', 
                        f'{col_letter_cards}167':'age_range_exceeded',
                        f'{col_letter_cards}169':'wants_to_visit_office', 
                        f'{col_letter_cards}170':'dissatisfied', 
                        f'{col_letter_cards}171':'not_applicable_unemployed', 
                        f'{col_letter_cards}172':'less_than_6_months', 
                        f'{col_letter_cards}173':'foreigner',
                        f'{col_letter_cards}174':'cash_only',
                        f'{col_letter_cards}175':'not_interested_loan', 
                        f'{col_letter_cards}176':'wants_amount_info',
                        f'{col_letter_cards}177':'banreservas_employee', 
                        f'{col_letter_cards}178':'credit_recovery', 
                        f'{col_letter_cards}179':'disconnected', 
                        f'{col_letter_cards}180':'very_low_limits'
                    }

                    for key,value in clasic_dispositions.items():
                        ws[key] = self.count_forms_by_state_and_cards(value,date,'Standard')

                    for key,value in gold_dispositions.items():
                        ws[key] = self.count_forms_by_state_and_cards(value,date,'Gold')

                    for key,value in multimoneda_dispositions.items():
                        ws[key] = self.count_forms_by_state_and_cards(value,date,'Multimoneda')

                    for key,value in platinum_dispositions.items():
                        ws[key] = self.count_forms_by_state_and_cards(value,date,'Platinum')
 
                    for key,value in multimoneda_gold_dispositions.items():
                        ws[key] = self.count_forms_by_state_and_cards(value,date,'Multimoneda Gold')

                
                    count += 1



            
            fp = BytesIO()
            wb.save(fp)
            fp.seek(0)

            excel_file = base64.b64encode(fp.read())

            fp.close()
            file_name = 'Reporte de Productividad.csv'
            attachment = self.env['ir.attachment'].create({
                'name': file_name,
                'type': 'binary',
                'datas': excel_file,
                'store_fname': file_name,
                'res_model': 'formulario.cliente',
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })
            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'self',
            }

